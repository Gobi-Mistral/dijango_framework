import csv
import time
from datetime import datetime
import logging
import platform
import openpyxl
import os
from pathlib import Path
from telnetlib import EC
from typing import Optional

from appium.webdriver.common.touch_action import TouchAction
from django.test import LiveServerTestCase
from appium import webdriver as appium_webdriver
from appium import webdriver as desktop_webdriver
from appium.webdriver.common.mobileby import MobileBy
from selenium import webdriver as selenium_webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait

from dijango_framework.helper import env

current_os = platform.system()
logger = logging.getLogger(__name__)

class TestBase(LiveServerTestCase):
    def __init__(self, *args, **kwargs):
        print("Test Base Init")
        super().__init__(*args, **kwargs)
        self.driver = None
        self.platform = env('PLATFORM')
        self.platform_name = env('PLATFORM_NAME')
        self.platform_version = env('PLATFORM_VER')
        self.device_name = env('DEVICE_NAME')
        self.device_app_name=env('DEVICE_APP_NAME')
        self.device_app_package = env('DEVICE_APP_PACKAGE')
        self.device_app_activity = env('DEVICE_APP_ACTIVITY')
        self.device_browser_name = env('DEVICE_BROWSER_NAME')
        self.appium_server = env('APPIUM_SERVER')
        self.web_url = env('WEB_URL')
        self.username= env('USERNAME')
        self.password = env('PASSWORD')
        self.wait_timeout = env('WAIT_TIMEOUT')
        self.browser = env('BROWSER')
        self.headless = env('HEADLESS')
        self.test_log_file = env('TEST_LOG_FILE')
        self.desktop_app = env('DESKTOP_APP')
        self.current_platform = current_os
        self.root_dir = Path(__file__).resolve().parent.parent.parent
        self.drivers_dir = str(self.root_dir) + "\\drivers\\"
        print("drivers_dir", self.drivers_dir)

    def setUp(self) -> None:
        print("Test Base Setup")
        if self.platform == "WEB":
            self.browser_configurations()

        elif self.platform == "ANDROID":
            self.device_configurations()

        elif self.platform == "DESKTOP":
            self.desktop_configuration()

    def tearDown(self) -> None:
        print("Test Base Teardown")
        self.driver.quit()

    ############################################################################
    # Browser & Device setup
    # These methods are unique. It can be used by given configuration.
    ############################################################################
    def browser_configurations(self):
        browser_options = None
        if self.browser == "CHROME":
            browser_options = selenium_webdriver.ChromeOptions()
        elif self.browser == "FIREFOX":
            browser_options = selenium_webdriver.FirefoxOptions()
        elif self.browser == "EDGE":
            browser_options = selenium_webdriver.EdgeOptions()

        if self.headless:
            if current_os == "Windows":
                browser_options.add_argument('--headless')
            elif current_os == "Linux":
                browser_options.add_argument('--no-sandbox')
                browser_options.add_argument('--headless')
                browser_options.add_argument('--disable-dev-shm-usage')

        if current_os == "Linux":
            browser_options.add_argument("window-size=1920,1080")

        browser_options.add_argument('--ignore-ssl-errors=yes')
        browser_options.add_argument('--ignore-certificate-errors')
        browser_options.add_argument("--start-maximized")

        if self.browser == "CHROME":
            driver_path = self.drivers_dir+"chromedriver.exe"
            self.driver = selenium_webdriver.Chrome(executable_path=driver_path, options=browser_options)
        elif self.browser == "FIREFOX":
            driver_path = self.drivers_dir+"geckodriver.exe"
            self.driver = selenium_webdriver.Firefox(executable_path=driver_path, options=browser_options)
        elif self.browser == "EDGE":
            driver_path = self.drivers_dir+"msedgedriver.exe"
            self.driver = selenium_webdriver.Edge(executable_path=driver_path, options=browser_options)

    def device_configurations(self):
        desired_caps = {}
        desired_caps['automationName'] = 'UiAutomator2'
        desired_caps['platformName'] = self.platform_name
        desired_caps['platformVersion'] = self.platform_version
        desired_caps['deviceName'] = self.platform_name
        desired_caps['udid'] = self.device_name
        desired_caps['autoGrantPermissions'] = True
        desired_caps['newCommandTimeout'] = 180  # 1 hour 15 minutes
        desired_caps['unicodeKeyboard'] = True
        # desired_caps['resetKeyboard'] = True

        if self.device_app_name == "Native":
            desired_caps['appPackage'] = self.device_app_package
            desired_caps['appActivity'] = self.device_app_activity

        elif self.device_app_name == "Hybrid":
            chromedriver_path = self.drivers_dir+"chromedriver.exe"
            desired_caps['browserName'] = self.device_browser_name
            desired_caps['chromedriverExecutable'] = chromedriver_path

        self.driver = appium_webdriver.Remote(self.appium_server, desired_caps)

    def desktop_configuration(self):
        desired_caps = {
            'platformName': 'Windows',
            'deviceName': 'WindowsPC',
            'app': self.desktop_app,
            'platformVersion': '10',
            'newCommandTimeout': 300  # Set the timeout (seconds) for new commands
        }

        self.driver = desktop_webdriver.Remote(
            command_executor=self.appium_server,  # WinAppDriver server address
            desired_capabilities=desired_caps
        )
        # self.driver.maximize_window()

    ############################################################################
    # Generic Methods for Web Browsers
    # These methods are generic. It can be used by any module on web based
    ############################################################################
    def get(self, relative_url: str) -> None:
        self.driver.get(relative_url)

    def click(self, by: str, selector: str) -> Optional[WebElement]:
        element = self.find(by, selector)
        try:
            element.click()
        except Exception as ex:
            self.driver.execute_script("arguments[0].click();", element)

    def _loop_through_selector_data_pair(self, *data: tuple):
        for by, selector, value in data:
            element = self.find(by, selector)
            yield element, value

    def find(self, by: str, selector: str) -> Optional[WebElement]:
        element = self.driver.find_element(by, selector)
        self.assertTrue(element, 'Element cannot be found: "%s"' % selector)
        return element

    def fill_form(self, *data: tuple, clear_input=True):
        for element, value in self._loop_through_selector_data_pair(*data):
            if clear_input:
                element.clear()
            element.send_keys(value)
            self.waitForSeconds(0.1)

    def click_form(self, *data: tuple):
        for selector, element in data:
            element = self.driver.find_element(selector, element)
            element.click()
            self.waitForSeconds(0.1)

    def select_dropdown_by_index(self, *data: tuple):
        for element, index in self._loop_through_selector_data_pair(*data):
            self.assertEqual(element.tag_name.lower(), 'select', 'Element is not a select box')
            Select(element).select_by_index(index)

    def select_dropdown_by_value(self, *data: tuple):
        for element, value in self._loop_through_selector_data_pair(*data):
            self.assertEqual(element.tag_name.lower(), 'select', 'Element is not a select box')
            Select(element).select_by_value(value)

    def select_dropdown_by_visible_text(self, *data: tuple):
        for element, text in self._loop_through_selector_data_pair(*data):
            self.assertEqual(element.tag_name.lower(), 'select', 'Element is not a select box')
            Select(element).select_by_visible_text(text)

    def click_element(self, by: str, element: str):
        self.driver.find_element(by, element).click()

    def type_text(self, by: str, element: str, value):
        self.driver.find_element(by, element).send_keys(value)

    def clear_and_type_text(self, by: str, element: str, value):
        self.driver.find_element(by, element).clear()
        self.waitForSeconds(1)
        self.driver.find_element(by, element).send_keys(value)

    def retrive_text(self, by: str, element: str):
        retrived_text = self.driver.find_element(by, element).text
        # print("Retrived Text -:- ", retrived_text)
        return retrived_text

    def verify_element_is_present(self, by: str, selector: str):
        element = self.driver.find_element(by, selector)
        element.is_displayed()

    def check_element_is_present(self, by: str, selector: str):
        status = False

        try:
            self.driver.find_element(by, selector)
            status = True
        except:
            status = False

        return status

    def move_to_element_by_text(self, text):
        element_visible = False
        max_scroll_count = 30
        scroll_count = 0
        while (not element_visible):
            try:
                self.driver.find_element_by_android_uiautomator(f"new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView(new UiSelector().text(\"{text}\"));")
                self.waitForSeconds(1)
                #print("*** INFO: Element found...")
                element_visible = True
            except:
                self.waitForSeconds(1)
                scroll_count += 1
                if scroll_count == max_scroll_count:
                    self.assertTrue(False, f"*** ERROR: Element {text} is not visible after waiting for {scroll_count} time")
                    break

    def move_to_element_by_id(self, id):
        element_visible = False
        max_scroll_count = 10
        scroll_count = 0
        while (not element_visible):
            try:
                self.driver.find_element_by_android_uiautomator( f"new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView(new UiSelector().resourceIdMatches(\"{id}\"));")
                self.waitForSeconds(2)
                # print("*** INFO: Element found...")
                element_visible = True
            except:
                self.waitForSeconds(1)
                scroll_count += 1
                if scroll_count == max_scroll_count:
                    self.assertTrue(False, f"*** ERROR: Element is not visible after waiting for {scroll_count} time")
                    break

    def verify_element_not_present(self, by: str, selector: str):
        variables = None
        try:
            self.driver.find_element(by, selector)
            variables = True
        except NoSuchElementException:
            variables = False
        self.assertFalse(variables, f"*** INFO: Element {selector} expected to be not present but it is")

    def waitForSeconds(self, Seconds):
        time.sleep(Seconds)

    def wait_until_visible(self, by: str, element: str, timeout: int = 30):
        element_visible = False
        elapsed_time = 0
        while (not element_visible):
            try:
                elements = WebDriverWait(self.driver, 1).until(
                    EC.visibility_of_element_located((by, element)))
                element_visible = True
            except:
                self.waitForSeconds(1)
                elapsed_time += 1
                if elapsed_time == timeout:
                    self.assertTrue(False, f"*** ERROR: Element is not visible after waiting for {elapsed_time} seconds")
                    break

    def wait_until_not_visible(self, by: str, element: str, timeout: int = 30):
        element_visible = False
        elapsed_time = 0
        while (not element_visible):
            try:
                elements = WebDriverWait(self.driver, 1).until(
                    EC.invisibility_of_element((by, element)))
                # print("*** INFO: Element found...")
                element_visible = True
            except:
                self.waitForSeconds(1)
                elapsed_time += 1
                if elapsed_time == timeout:
                    self.assertTrue(False, f"*** ERROR: Element is not visible after waiting for {elapsed_time} seconds")
                    break

    def wait_until_presence(self, by: str, element: str, timeout: int = 30):
        element_visible = False
        elapsed_time = 0
        while (not element_visible):
            try:
                elements = WebDriverWait(self.driver, 1).until(
                    EC.presence_of_element_located((by, element)))
                element_visible = True
            except:
                self.waitForSeconds(1)
                elapsed_time += 1
                if elapsed_time == timeout:
                    self.assertTrue(False, f"*** ERROR: Element is not present after waiting for {elapsed_time} seconds")
                    break

    def wait_until_element_to_be_selected(self, by: str, element: str, timeout: int = 30):
        element_visible = False
        elapsed_time = 0
        while (not element_visible):
            try:
                elements = WebDriverWait(self.driver, 1).until(
                    EC.element_located_to_be_selected((by, element)))
                element_visible = True
            except:
                self.waitForSeconds(1)
                elapsed_time += 1
                if elapsed_time == timeout:
                    self.assertTrue(False, f"*** ERROR: Element is not selected after waiting for {elapsed_time} seconds")
                    break

    def wait_until_element_to_be_clickable(self, by: str, element: str, timeout: int = 30):
        element_visible = False
        elapsed_time = 0
        while (not element_visible):
            try:
                elements = WebDriverWait(self.driver, 1).until(
                    EC.element_to_be_clickable((by, element)))
                element_visible = True
            except:
                self.waitForSeconds(1)
                elapsed_time += 1
                if elapsed_time == timeout:
                    self.assertTrue(False, f"*** ERROR: Element is not clickable after waiting for {elapsed_time} seconds")
                    break

    def alert_accept(self):
        alert_msg = None
        try:
            self.waitForSeconds(1)
            WebDriverWait(self.driver, 30).until(EC.alert_is_present(), 'Timed out waiting for a popup to be appear.')
            alert = self.driver.switch_to.alert
            alert_msg = alert.text
            alert.accept()
            print("Alert is Present -:- ", alert_msg)
        except:
            pass
            print("Alert is Not Present")

        self.waitForSeconds(1)

    def alert_dismiss(self):
        alert_msg = None
        try:
            self.waitForSeconds(1)
            WebDriverWait(self.driver, 30).until(EC.alert_is_present(), 'Timed out waiting for a popup to be appear.')
            alert = self.driver.switch_to.alert
            alert_msg = alert.text
            alert.dismiss()
            print("Alert is Present -:- ", alert_msg)
        except:
            pass
            print("Alert is Not Present")

        self.waitForSeconds(1)

    def take_screen_shot(self, file_name, dir="screenshot_failures"):
        now = datetime.now()
        filename = file_name+"_"+now.strftime("%d%m%Y_%H%M%S")+".png"
        name = filename.replace(" ", "_")
        path = f"{dir}/{name}"
        self.driver.save_screenshot(path)
        return filename

    def log_test_result(self, page_title, test_id, log_message):
        current_time = datetime.utcnow()
        print(f'{current_time}, {page_title}, {test_id} {log_message}')
        content = ','.join(list([
            '{date:%Y-%m-%d %H:%M:%S}'.format(date=datetime.utcnow()),
            page_title,
            str(test_id),
            log_message,
        ]))
        with open(self.test_log_file, 'a') as f:
            f.write(content + '\n')

        if log_message.__eq__("FAIL"):
            name = page_title+"_"+test_id+".png"
            filename = name.replace(" ", "_")
            self.take_screen_shot(filename)

    def scrollDown(self):
        handle_one_size = self.driver.get_window_size()
        scroll_start = handle_one_size['height'] * 0.5
        scroll_end = handle_one_size['height'] * 0.2

        action = TouchAction(self.driver)
        for i in range(10):
            action.press(x=0, y=scroll_start).move_to(x=0, y=scroll_end).release().perform()
            self.waitForSeconds(1)

    def scroll_down_until_visible(self, by: str, element: str):
        element_visible = False
        max_scroll_count = 100
        scroll_count = 0

        handle_one_size = self.driver.get_window_size()
        scroll_x = handle_one_size['width'] * 0.5
        scroll_start = handle_one_size['height'] * 0.5
        scroll_end = handle_one_size['height'] * 0.2

        action = TouchAction(self.driver)

        while (not element_visible):
            try:
                self.driver.find_element(by, element)
                # print("*** INFO: Element found...")
                element_visible = True
            except:
                action.press(x=scroll_x, y=scroll_start).move_to(x=scroll_x, y=scroll_end).release().perform()
                self.waitForSeconds(1)
                scroll_count += 1
                if scroll_count == max_scroll_count:
                    self.assertTrue(False, f"*** ERROR: Element is not visible after scrolling for {scroll_count} time")
                    break

    def scroll_up_until_visible(self, by: str, element: str):
        element_visible = False
        max_scroll_count = 100
        scroll_count = 0

        handle_one_size = self.driver.get_window_size()
        scroll_x = handle_one_size['width'] * 0.5
        scroll_start = handle_one_size['height'] * 0.2
        scroll_end = handle_one_size['height'] * 0.5

        action = TouchAction(self.driver)

        while (not element_visible):
            try:
                self.driver.find_element(by, element)
                # print("*** INFO: Element found...")
                element_visible = True
            except:
                action.press(x=scroll_x, y=scroll_start).move_to(x=scroll_x, y=scroll_end).release().perform()
                self.waitForSeconds(1)
                scroll_count += 1
                if scroll_count == max_scroll_count:
                    self.assertTrue(False, f"*** ERROR: Element is not visible after scrolling for {scroll_count} time")
                    break

    def scroll_down_web_page(self):
        self.driver.execute_script("window.scrollTo(0, 150)")

    def highlight_the_element(self, element: str):
        highlight_script = "var element = arguments[0]; element.style.border = '2px solid red'; element.style.backgroundColor = 'yellow';"
        self.driver.execute_script(highlight_script, element)

    def reset_the_highlighted_element(self, element: str):
        reset_highlight_script = "var element = arguments[0]; element.style.border = ''; element.style.backgroundColor = '';"
        self.driver.execute_script(reset_highlight_script, element)

    def navigate_back(self):
        self.driver.back() # Perform a back navigation
        self.waitForSeconds(2)

    def navigate_forward(self):
        self.driver.forward() # Perform a forward navigation
        self.waitForSeconds(2)

    def page_refresh(self):
        self.driver.refresh() # Refresh the page
        self.waitForSeconds(2)

    def execute_java_scripts(self, scripts: str):
        # Execute JavaScript code
        self.driver.execute_script(scripts)

    def delete_all_cookies(self):
        self.driver.delete_all_cookies()

    def maximize_window(self):
        self.driver.maximize_window() # maximize window position

    def minimize_window(self):
        self.driver.minimize_window() # minimize window position

    def fullscreen_window(self):
        self.driver.fullscreen_window() # Full Screen window

    def set_window_size(self, size_x, size_y):
        self.driver.set_window_size(size_x, size_y)

    def set_page_load_timeout(self, seconds_To_wait: int):
        self.driver.set_page_load_timeout(seconds_To_wait)

    def retrieve_current_url(self):
        url = self.driver.current_url
        print("Page Current Url -:- ", url)
        return url

    def click_and_hold(self, element: str):
        action = ActionChains(self.driver)
        action.click_and_hold(on_element=element).perform()

    def double_click(self, element: str):
        action = ActionChains(self.driver)
        action.double_click(on_element=element).perform()

    def drag_and_drop(self, source_element: str, target_element: str):
        action = ActionChains(self.driver)
        action.drag_and_drop(source_element, target_element).perform()

    def move_to_element(self, element: str):
        action = ActionChains(self.driver)
        action.move_to_element(element).perform()

    def open_new_tab(self, url: str):
        self.driver.execute_script(f"window.open('{url}')")

    def clear_input_field(self, by: str, element: str):
        self.driver.find_element(by, element).clear()

    def get_page_source(self):
        page_source = self.driver.page_source
        return page_source

    def implicitly_wait(self, seconds: int):
        self.driver.implicitly_wait(seconds) # in seconds

    def get_attribute(self, by: str, element: str, value: str):
        element = self.driver.find_element(by, element)
        get_values = element.get_attribute(value)
        return get_values

    def switch_to_new_window(self):
        current_window_handle = self.driver.current_window_handle

        for handle in self.driver.window_handles: # Handling multiple windows and activate recent window
            if handle != current_window_handle:
                self.driver.switch_to.window(handle)
                break

    def switch_to_frame(self, by: str, element: str):
        element = self.driver.find_element(by, element)
        self.driver.switch_to.frame(element)

    def switch_to_default_content(self):
        self.driver.switch_to.default_content()

    def verify_element_is_selected(self, by: str, element: str):
        status = None
        find_element = self.driver.find_element(by, element)
        if find_element.is_selected():
            status = True
            print("Element is selected")
        else:
            status = False
            print("Element is not selected")

        return status

    def verify_element_is_enabled(self, by: str, element: str):
        status = None
        find_element = self.driver.find_element(by, element)
        if find_element.is_enabled():
            status = True
            print("Element is enabled")
        else:
            status = False
            print("Element is not enabled")

        return status

    def submit(self,by:str, element:str):
        find_element = self.driver.find_element(by, element)
        find_element.submit()

    def close_tab(self):
        self.driver.close()

    def jsClick(self, by:str, element: str):
        find_element = self.driver.find_element(by, element)
        self.driver.execute_script("arguments[0].click();", find_element) # Use JavaScript to perform the click

    def get_title(self):
        actual_title = self.driver.title
        return actual_title

    ############################################################################
    # Python Generic Methods
    ############################################################################
    def read_text_file(self, file_name):
        file_name = "D:/work-space/digi-development/dijango_framework/requirements.txt"

        with open(file_name, "r") as file:
            # data = file.read()
            for line in file:
                print("Read Text File Line By Line -:- ", line.strip())

    def read_csv_file(self, file_name):
        file_name = "D:/work-space/digi-development/dijango_framework/test_log.csv"

        with open(file_name, "r") as csvfile:
            csv_reader = csv.reader(csvfile)

            for row in csv_reader:
                print("Read CSV File Row By Row -:- ", row)

    def read_excel_file(self, file_name):
        file_name = "D:/work-space/digi-development/dijango_framework/Test_Report.xlsx"

        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active

        for row in worksheet.iter_rows(values_only=True):
            print("Read Excel File By Row -:- ", row)

        for row in worksheet.iter_rows():
            for cell in row:
                print("Read Excel File By Cell -:- ", cell.value)

    ############################################################################
    # Generic Methods for Android devices
    # These methods are generic. It can be used by any module on android device based
    ############################################################################
    def launch_activity(self, packageName, activity):
        # Launching the Android File Manager
        self.driver.start_activity(packageName, activity)
        self.waitForSeconds(3)

    def close_activity(self, packageName):
        # Closing the Android File Manager
        self.driver.terminate_app(packageName)
        self.waitForSeconds(1)

    ############################################################################
    # Generic Methods for Desktop applications
    # These methods are generic. It can be used by any module on desktop devices
    ############################################################################
    def basic(self):
        desired_caps = {}
        self.driver = appium_webdriver.Remote(
            command_executor=self.appium_server,  # WinAppDriver server address
            desired_capabilities=desired_caps
        )